import openpyxl

inv_file = openpyxl.load_workbook("companyInventory.xlsx")
product_list = inv_file["Inventory"]

products_per_supplier = {}
total_inventory_value_per_supplier = {}
products_low_inventory = {}
restock_threshold = 150

for product_row in range(2, product_list.max_row + 1):
    product_num = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    unit_price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory_value = product_list.cell(product_row, 5)

    # Calculate number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_product = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_product + 1
    else:
        products_per_supplier[supplier_name] = 1

    # Calculate total value of inventory per supplier
    if supplier_name in total_inventory_value_per_supplier:
        current_inventory_value = total_inventory_value_per_supplier.get(supplier_name)
        total_inventory_value_per_supplier[supplier_name] = current_inventory_value + inventory * unit_price
    else:
        total_inventory_value_per_supplier[supplier_name] = inventory * unit_price

    # Find low inventory product by restock threshold
    if inventory <= restock_threshold:
        products_low_inventory[product_num] = inventory
        print(f"Product {product_num} inventory is running low at {inventory}, less than {restock_threshold} needs "
              f"restock!")

    # Calculate and add inventory value per product
    inventory_value.value = inventory * unit_price

print(products_low_inventory)
print(products_per_supplier)
print(total_inventory_value_per_supplier)

inv_file.save("companyInventory_with_value.xlsx")
